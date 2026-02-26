"""SAT and CNE report generation module."""
import os
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from database import db, Station, FuelTransaction, InventorySnapshot, Report


REPORT_DIR = os.path.join(os.path.dirname(__file__), "generated_reports")
os.makedirs(REPORT_DIR, exist_ok=True)

FUEL_LABELS = {"magna": "Magna (Regular)", "premium": "Premium", "diesel": "Diesel"}


def get_daily_summary(station_id, fuel_type, target_date):
    """Get received, sold, and closing inventory for a station/fuel/date."""
    start = datetime.combine(target_date, datetime.min.time())
    end = datetime.combine(target_date, datetime.max.time())

    received = db.session.query(db.func.coalesce(db.func.sum(FuelTransaction.liters), 0)).filter(
        FuelTransaction.station_id == station_id,
        FuelTransaction.fuel_type == fuel_type,
        FuelTransaction.transaction_type == "received",
        FuelTransaction.timestamp.between(start, end),
    ).scalar()

    sold = db.session.query(db.func.coalesce(db.func.sum(FuelTransaction.liters), 0)).filter(
        FuelTransaction.station_id == station_id,
        FuelTransaction.fuel_type == fuel_type,
        FuelTransaction.transaction_type == "sold",
        FuelTransaction.timestamp.between(start, end),
    ).scalar()

    snapshot = InventorySnapshot.query.filter_by(
        station_id=station_id, fuel_type=fuel_type, snapshot_date=target_date
    ).first()

    closing = snapshot.liters_on_hand if snapshot else 0

    return {
        "received": float(received),
        "sold": float(sold),
        "closing": float(closing),
    }


def generate_sat_volumetric(target_date=None):
    """Generate SAT volumetric control daily report for all stations."""
    if target_date is None:
        target_date = date.today()

    stations = Station.query.filter_by(active=True).order_by(Station.code).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Control Volumetrico SAT"

    # Header styling
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="0A1628", end_color="0A1628", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title rows
    ws.merge_cells("A1:J1")
    ws["A1"] = "CONTROL VOLUMETRICO DIARIO - SAT"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="0A1628")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Fecha: {target_date.strftime('%d/%m/%Y')} | Generado por Control Petro"
    ws["A2"].font = Font(name="Arial", size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers
    headers = [
        "Estacion", "Codigo", "Combustible",
        "Inventario Inicial (L)", "Litros Recibidos",
        "Litros Vendidos", "Inventario Final (L)",
        "Capacidad Tanque (L)", "% Ocupacion", "Estado"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    widths = [25, 10, 18, 18, 16, 16, 18, 18, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 5
    total_received = total_sold = 0
    status_fill = {
        "Normal": PatternFill(start_color="E8F5E9", fill_type="solid"),
        "Bajo": PatternFill(start_color="FFF3E0", fill_type="solid"),
        "Critico": PatternFill(start_color="FFEBEE", fill_type="solid"),
    }

    for station in stations:
        for fuel_type in ["magna", "premium", "diesel"]:
            summary = get_daily_summary(station.id, fuel_type, target_date)
            capacity = getattr(station, f"{fuel_type}_capacity", 40000)
            pct = (summary["closing"] / capacity * 100) if capacity > 0 else 0
            opening = summary["closing"] - summary["received"] + summary["sold"]

            status = "Normal" if pct > 40 else ("Bajo" if pct > 25 else "Critico")

            values = [
                station.name, station.code, FUEL_LABELS.get(fuel_type, fuel_type),
                round(opening, 1), round(summary["received"], 1),
                round(summary["sold"], 1), round(summary["closing"], 1),
                capacity, f"{pct:.1f}%", status,
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if col > 2 else "left")
                if col == 10:
                    cell.fill = status_fill.get(status, PatternFill())

            total_received += summary["received"]
            total_sold += summary["sold"]
            row += 1

    # Totals row
    row += 1
    ws.cell(row=row, column=1, value="TOTALES").font = Font(bold=True, size=11)
    ws.cell(row=row, column=5, value=round(total_received, 1)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=round(total_sold, 1)).font = Font(bold=True)

    filename = f"SAT_Volumetrico_{target_date.strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(REPORT_DIR, filename)
    wb.save(filepath)

    # Record in database
    report = Report(
        report_type="sat_volumetric",
        report_date=target_date,
        status="generated",
        file_path=filepath,
        created_at=datetime.utcnow(),
        details=f"{len(stations)} estaciones, {total_received:.0f}L recibidos, {total_sold:.0f}L vendidos",
    )
    db.session.add(report)
    db.session.commit()

    return {"filename": filename, "filepath": filepath, "report_id": report.id,
            "stations": len(stations), "total_received": total_received, "total_sold": total_sold}


def generate_cne_weekly(target_date=None):
    """Generate CNE weekly report with prices, volumes, and quality data."""
    if target_date is None:
        target_date = date.today()

    week_start = target_date - timedelta(days=target_date.weekday())
    week_end = week_start + timedelta(days=6)

    stations = Station.query.filter_by(active=True).order_by(Station.code).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Semanal CNE"

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:I1")
    ws["A1"] = "REPORTE SEMANAL - COMISION NACIONAL DE ENERGIA (CNE)"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="0A1628")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Semana: {week_start.strftime('%d/%m/%Y')} - {week_end.strftime('%d/%m/%Y')} | Generado por Control Petro"
    ws["A2"].font = Font(name="Arial", size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = [
        "Estacion", "Codigo", "Combustible",
        "Volumen Semanal (L)", "Precio Promedio ($/L)",
        "Ingresos ($)", "Calidad NOM-016", "Quejas Clientes", "Estado Permiso"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    widths = [25, 10, 18, 18, 18, 18, 16, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    prices = {"magna": 23.45, "premium": 25.12, "diesel": 24.78}
    row = 5
    for station in stations:
        for fuel_type in ["magna", "premium", "diesel"]:
            start_dt = datetime.combine(week_start, datetime.min.time())
            end_dt = datetime.combine(week_end, datetime.max.time())
            weekly_sold = db.session.query(
                db.func.coalesce(db.func.sum(FuelTransaction.liters), 0)
            ).filter(
                FuelTransaction.station_id == station.id,
                FuelTransaction.fuel_type == fuel_type,
                FuelTransaction.transaction_type == "sold",
                FuelTransaction.timestamp.between(start_dt, end_dt),
            ).scalar()
            weekly_sold = float(weekly_sold)
            price = prices.get(fuel_type, 24.0)
            revenue = weekly_sold * price

            values = [
                station.name, station.code, FUEL_LABELS.get(fuel_type, fuel_type),
                round(weekly_sold, 1), f"${price:.2f}",
                f"${revenue:,.2f}", "Cumple", 0, "Vigente",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if col > 2 else "left")
            row += 1

    filename = f"CNE_Semanal_{week_start.strftime('%Y%m%d')}_{week_end.strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(REPORT_DIR, filename)
    wb.save(filepath)

    report = Report(
        report_type="cne_weekly",
        report_date=target_date,
        status="generated",
        file_path=filepath,
        created_at=datetime.utcnow(),
        details=f"Semana {week_start.strftime('%d/%m')} - {week_end.strftime('%d/%m')}, {len(stations)} estaciones",
    )
    db.session.add(report)
    db.session.commit()

    return {"filename": filename, "filepath": filepath, "report_id": report.id}


def generate_inventory_close(target_date=None):
    """Generate end-of-day inventory closing report."""
    if target_date is None:
        target_date = date.today()

    stations = Station.query.filter_by(active=True).order_by(Station.code).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario de Cierre"

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:H1")
    ws["A1"] = "INVENTARIO DE CIERRE DIARIO"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="0A1628")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Fecha: {target_date.strftime('%d/%m/%Y')} | Generado por Control Petro"
    ws["A2"].font = Font(name="Arial", size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = [
        "Estacion", "Codigo", "Magna (L)", "Premium (L)",
        "Diesel (L)", "Total (L)", "Capacidad Total (L)", "% Promedio"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    widths = [25, 10, 16, 16, 16, 16, 18, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 5
    for station in stations:
        levels = {}
        for ft in ["magna", "premium", "diesel"]:
            snap = InventorySnapshot.query.filter_by(
                station_id=station.id, fuel_type=ft, snapshot_date=target_date
            ).first()
            levels[ft] = snap.liters_on_hand if snap else 0

        total_inv = sum(levels.values())
        total_cap = station.magna_capacity + station.premium_capacity + station.diesel_capacity
        avg_pct = (total_inv / total_cap * 100) if total_cap > 0 else 0

        values = [
            station.name, station.code,
            round(levels["magna"], 1), round(levels["premium"], 1),
            round(levels["diesel"], 1), round(total_inv, 1),
            total_cap, f"{avg_pct:.1f}%"
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col > 2 else "left")
        row += 1

    filename = f"Inventario_Cierre_{target_date.strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(REPORT_DIR, filename)
    wb.save(filepath)

    report = Report(
        report_type="inventory_close",
        report_date=target_date,
        status="generated",
        file_path=filepath,
        created_at=datetime.utcnow(),
        details=f"{len(stations)} estaciones",
    )
    db.session.add(report)
    db.session.commit()

    return {"filename": filename, "filepath": filepath, "report_id": report.id}


def generate_price_report(target_date=None):
    """Generate price and tariff report."""
    if target_date is None:
        target_date = date.today()

    stations = Station.query.filter_by(active=True).order_by(Station.code).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Precios y Tarifas"

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:G1")
    ws["A1"] = "REPORTE DE PRECIOS Y TARIFAS"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="0A1628")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Fecha: {target_date.strftime('%d/%m/%Y')} | Generado por Control Petro"
    ws["A2"].font = Font(name="Arial", size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["Estacion", "Codigo", "Magna ($/L)", "Premium ($/L)", "Diesel ($/L)", "Descuentos", "Observaciones"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    row = 5
    for station in stations:
        values = [station.name, station.code, "$23.45", "$25.12", "$24.78", "N/A", "Precio vigente"]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col > 2 else "left")
        row += 1

    filename = f"Precios_Tarifas_{target_date.strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(REPORT_DIR, filename)
    wb.save(filepath)

    report = Report(
        report_type="price_tariff",
        report_date=target_date,
        status="generated",
        file_path=filepath,
        created_at=datetime.utcnow(),
    )
    db.session.add(report)
    db.session.commit()

    return {"filename": filename, "filepath": filepath, "report_id": report.id}


def get_report_history(limit=50):
    """Get recent report generation history."""
    reports = Report.query.order_by(Report.created_at.desc()).limit(limit).all()
    return [{
        "id": r.id,
        "type": r.report_type,
        "date": r.report_date.isoformat(),
        "status": r.status,
        "file_path": r.file_path,
        "sent_at": r.sent_at.isoformat() if r.sent_at else None,
        "created_at": r.created_at.isoformat(),
        "details": r.details,
    } for r in reports]


def mark_report_sent(report_id):
    """Mark a report as sent."""
    report = Report.query.get(report_id)
    if report:
        report.status = "sent"
        report.sent_at = datetime.utcnow()
        db.session.commit()
        return True
    return False
